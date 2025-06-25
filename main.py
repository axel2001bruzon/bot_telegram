import os
import re
import pandas as pd
from telegram import (
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    InputFile,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

TOKEN = "7967825027:AAHCvsCZ2_g9t7EWA-I8tLd9gtibFfdHVkg"

def build_main_menu_inline():
    buttons = [
        [InlineKeyboardButton("Filter Telegram Groups", callback_data="filter_telegram")],
        [InlineKeyboardButton("Filter WhatsApp Groups", callback_data="filter_whatsapp")],
        [InlineKeyboardButton("Filter Files", callback_data="filter_files")]
    ]
    return InlineKeyboardMarkup(buttons)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Welcome. Please choose an option from the menu:",
        reply_markup=build_main_menu_inline()
    )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    context.user_data.clear()
    if data == "filter_telegram":
        context.user_data["expecting_telegram_links"] = True
        await query.edit_message_text(
            "📥 You selected *Filter Telegram Groups*\n"
            "Send me Telegram group links (example: https://t.me/cubanjobs), and I will remove duplicates.",
            parse_mode="Markdown"
        )
    elif data == "filter_whatsapp":
        context.user_data["expecting_whatsapp_links"] = True
        await query.edit_message_text(
            "📥 You selected *Filter WhatsApp Groups*\n"
            "Send me WhatsApp group links, and I will remove duplicates.",
            parse_mode="Markdown"
        )
    elif data == "filter_files":
        context.user_data["expecting_file_upload"] = True
        await query.edit_message_text(
            "📁 You selected *Filter Files*\nSend me a `.xlsx` or `.txt` file containing links.",
            parse_mode="Markdown"
        )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    # Direct link filtering
    if context.user_data.get("expecting_whatsapp_links") or context.user_data.get("expecting_telegram_links"):
        whatsapp_pattern = r'https://chat\.whatsapp\.com/[A-Za-z0-9]+'
        telegram_pattern = r'https://t\.me/(?:joinchat/|s/|\+)?[A-Za-z0-9_]+'

        pattern = whatsapp_pattern if context.user_data.get("expecting_whatsapp_links") else telegram_pattern
        links = re.findall(pattern, text)

        if not links and context.user_data.get("expecting_telegram_links"):
            # Search simple telegram links if none found
            links = re.findall(r'https://t\.me/[A-Za-z0-9_]+', text)

        if links:
            unique_links = list(dict.fromkeys(links))
            total = len(links)
            duplicates = total - len(unique_links)
            context.user_data["filtered_links"] = unique_links
            context.user_data["waiting_for_file_name"] = True
            context.user_data["expecting_whatsapp_links"] = False
            context.user_data["expecting_telegram_links"] = False

            enumerated = "\n".join(f"{i+1}- {l}" for i, l in enumerate(unique_links))
            await update.message.reply_text(
                f"✅ *Links processed:*\n🔢 Total received: {total}\n♻️ Duplicates removed: {duplicates}\n\n📤 *Unique links:*\n{enumerated}",
                parse_mode="Markdown"
            )
            await update.message.reply_text("📝 Please type the filename for the `.xlsx` file (without extension):")
        else:
            await update.message.reply_text("❗ No valid links found.")
        return

    # Save links to Excel file
    if context.user_data.get("waiting_for_file_name"):
        file_base = re.sub(r'[^\w\-]', '_', text)
        file_name = f"{file_base}.xlsx"
        links = context.user_data.get("filtered_links", [])

        if not links:
            await update.message.reply_text("❗ There are no links to save.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Links"

        ws["A1"] = "No."
        ws["B1"] = "Link"
        ws["A1"].font = ws["B1"].font = Font(bold=True)

        for i, link in enumerate(links, 1):
            ws.cell(row=i+1, column=1, value=i)
            cell = ws.cell(row=i+1, column=2, value=link)
            cell.hyperlink = link
            cell.style = "Hyperlink"

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_name)
        wb.close()
        with open(file_name, "rb") as f:
            await update.message.reply_document(InputFile(f, filename=file_name), caption="📄 Excel file generated successfully.")
        os.remove(file_name)
        context.user_data.clear()
        return

    await update.message.reply_text("❗ Please choose an option from the menu first.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("expecting_file_upload"):
        return

    file = update.message.document
    original_file_name = file.file_name  # Keep original name
    file_path = f"temp_{original_file_name}"
    file_obj = await file.get_file()
    await file_obj.download_to_drive(file_path)

    all_links = set()  # Set to remove duplicates globally
    filtered_data = {}

    try:
        if original_file_name.endswith(".xlsx"):
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                filtered_data[sheet_name] = []  # Keep original structure

                columns = [[] for _ in range(ws.max_column)]  # List to store columns without empty spaces

                for row in ws.iter_rows(values_only=True):
                    for col_idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            matches = re.findall(
                                r'https://(?:chat\.whatsapp\.com/[A-Za-z0-9]+|t\.me/[A-Za-z0-9_+/]+)', cell
                            )
                            filtered_links = [link for link in matches if link not in all_links]
                            all_links.update(filtered_links)
                            if filtered_links:
                                columns[col_idx].extend(filtered_links)

                # Rebuild rows without empty spaces
                max_rows = max(len(col) for col in columns)
                for i in range(max_rows):
                    new_row = [columns[col_idx][i] if i < len(columns[col_idx]) else None for col_idx in range(len(columns))]
                    filtered_data[sheet_name].append(new_row)

            wb.close()

        elif original_file_name.endswith(".txt"):
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
                matches = re.findall(
                    r'https://(?:chat\.whatsapp\.com/[A-Za-z0-9]+|t\.me/[A-Za-z0-9_+/]+)', content
                )
                all_links.update(matches)

        else:
            await update.message.reply_text("❗ Only `.xlsx` or `.txt` files are accepted.")
            return

        if not all_links:
            await update.message.reply_text("❗ No links found in the file.")
            return

        total_links_before = sum(len(row) for sheet in filtered_data.values() for row in sheet)  # Count before removing duplicates
        unique_links = list(dict.fromkeys(all_links))  # Remove duplicates globally
        total_links_after = len(unique_links)  # Final count without duplicates
        duplicates = total_links_before - total_links_after  # Number of duplicates removed

        await update.message.reply_text(
            f"✅ *Links processed:*\n🔢 Total received: {total_links_before}\n♻️ Duplicates removed: {duplicates}",
            parse_mode="Markdown"
        )

        output_file = f"filtered_{original_file_name}"  # Keep original name

        # Save with pandas ExcelWriter using xlsxwriter engine
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            for sheet_name, rows in filtered_data.items():
                df = pd.DataFrame(rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Adjust column widths
                worksheet.set_column(0, len(rows[0]) - 1, 70)

        # Send the file with modified name
        with open(output_file, "rb") as f:
            await update.message.reply_document(
                InputFile(f, filename=output_file), caption="📤 Filtered file generated successfully."
            )
        os.remove(output_file)

    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        context.user_data.clear()

async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❓ Unknown command. Please use /start to begin.")

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.COMMAND, unknown_command))

    print("Bot started...")
    app.run_polling()

if __name__ == "__main__":
    main()
































